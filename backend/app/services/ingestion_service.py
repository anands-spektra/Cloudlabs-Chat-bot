import io


async def extract_text(content: bytes, filename: str, content_type: str) -> str | None:
    try:
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

        if content_type in ("text/markdown", "text/x-markdown") or ext == "md":
            return content.decode("utf-8", errors="replace")

        if ext == "txt" or content_type == "text/plain":
            return content.decode("utf-8", errors="replace")

        if content_type == "application/pdf" or ext == "pdf":
            return _extract_pdf(content)

        if content_type in (
            "application/msword",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ) or ext in ("doc", "docx"):
            return _extract_docx(content)

        if content_type in (
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ) or ext in ("xls", "xlsx"):
            return _extract_xlsx(content)

    except Exception:
        return None
    return None


def _extract_pdf(content: bytes) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(content))
        parts = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                parts.append(text)
        return "\n".join(parts)
    except Exception:
        return ""


def _extract_docx(content: bytes) -> str:
    try:
        from docx import Document
        doc = Document(io.BytesIO(content))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception:
        return ""


def _extract_xlsx(content: bytes) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"## Sheet: {sheet_name}")
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            # Use first non-empty row as header if it looks like one
            headers = [str(c) if c is not None else "" for c in rows[0]]
            parts.append(" | ".join(headers))
            for row in rows[1:]:
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    parts.append(" | ".join(cells))
        wb.close()
        return "\n".join(parts)
    except Exception:
        return ""
